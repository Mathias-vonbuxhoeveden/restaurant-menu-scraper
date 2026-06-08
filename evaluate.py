#!/usr/bin/env python3
"""
evaluate.py — Evaluate menu scraper accuracy across multiple restaurants

Usage:
  python evaluate.py --scraped Ricordi.xlsx --ground-truth ground_truth.xlsx
"""

import argparse
import difflib
import json
import re
import sys

import openpyxl
from pydantic import BaseModel


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

class MatchedItem(BaseModel):
    dish: str
    price: str
    category: str
    description: str
    desc_ok: bool

class WrongFieldsItem(BaseModel):
    dish: str
    price_ok: bool
    category_ok: bool
    desc_ok: bool
    scraped_price: str
    expected_price: str
    scraped_category: str
    expected_category: str
    scraped_description: str
    expected_description: str

class EvaluationResult(BaseModel):
    matched:      list[MatchedItem]
    wrong_fields: list[WrongFieldsItem]
    missing:      list[str]
    hallucinated: list[str]
    precision: float
    recall:    float
    name_accuracy:        float
    price_accuracy:       float
    category_accuracy:    float
    description_accuracy: float


# ---------------------------------------------------------------------------
# File reading
# ---------------------------------------------------------------------------

def read_xlsx_all_sheets(path: str) -> dict[str, list[dict]]:
    """Returns {sheet_name: [rows]} for every sheet in the workbook."""
    wb = openpyxl.load_workbook(path)
    sheets: dict[str, list[dict]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[dict] = []
        headers: list[str] | None = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c).strip() if c is not None else "" for c in row]
                continue
            if any(c is not None for c in row):
                rows.append({
                    h: (str(c).strip() if c is not None else "")
                    for h, c in zip(headers, row)
                })
        sheets[sheet_name] = rows
    return sheets


# ---------------------------------------------------------------------------
# Normalisation
# ---------------------------------------------------------------------------

_VEGE_RE = re.compile(r"\s*\(V\+?\)\s*", re.IGNORECASE)


def normalize_name(name: str) -> str:
    return re.sub(r"[''`']", "", name).strip().lower()

def normalize_price(price: str) -> str:
    """Normalize price to a plain integer string, e.g. '115.0 kr' → '115'."""
    price = price.strip()
    if not price:
        return ""
    try:
        return str(int(float(re.sub(r"[^\d.,]", "", price).replace(",", "."))))
    except (ValueError, TypeError):
        return re.sub(r"\D", "", price)

def normalize_category(cat: str) -> str:
    return cat.strip().lower()

def normalize_desc(desc: str) -> str:
    return _VEGE_RE.sub(" ", desc).strip().lower()

def _fuzzy_ok(a: str, b: str, threshold: float = 0.8) -> bool:
    if not a and not b:
        return True
    if not a or not b:
        return False
    return difflib.SequenceMatcher(None, a, b).ratio() >= threshold

def category_similar(a: str, b: str) -> bool:
    na, nb = normalize_category(a), normalize_category(b)
    if not na and not nb:
        return True
    if not na or not nb:
        return False
    return name_similarity(na, nb) >= FUZZY_NAME_THRESHOLD

def desc_similar(a: str, b: str) -> bool:
    return _fuzzy_ok(normalize_desc(a), normalize_desc(b))


FUZZY_NAME_THRESHOLD = 0.85


def name_similarity(a: str, b: str) -> float:
    if a == b:
        return 1.0
    set_a, set_b = set(a.split()), set(b.split())
    shorter = set_a if len(set_a) <= len(set_b) else set_b
    longer  = set_b if len(set_a) <= len(set_b) else set_a
    token_score = len(shorter & longer) / len(shorter) if shorter else 0.0
    char_score  = difflib.SequenceMatcher(None, a, b).ratio()
    return max(token_score, char_score)

def find_scraped_key(gt_key: str, scraped_map: dict[str, list], matched_indices: set[tuple[str, int]]) -> tuple[str | None, int]:
    """Returns (key, index) into scraped_map[key], or (None, 0) if no match."""
    # Exact match — pick first unmatched entry
    if gt_key in scraped_map:
        for i in range(len(scraped_map[gt_key])):
            if (gt_key, i) not in matched_indices:
                return gt_key, i
    # Fuzzy match
    best_key, best_score, best_idx = None, 0.0, 0
    for key, entries in scraped_map.items():
        score = name_similarity(gt_key, key)
        if score > best_score:
            for i in range(len(entries)):
                if (key, i) not in matched_indices:
                    best_score, best_key, best_idx = score, key, i
                    break
    if best_key and best_score >= FUZZY_NAME_THRESHOLD:
        return best_key, best_idx
    return None, 0

def get_row_fields(row: dict) -> tuple[str, str, str, str]:
    dish        = row.get("Name")        or row.get("name")        or row.get("Rätt")        or row.get("rätt")        or row.get("dish")        or list(row.values())[0]
    price       = row.get("Price")       or row.get("price")       or row.get("Pris")        or row.get("pris")        or (list(row.values())[1] if len(row) > 1 else "")
    category    = row.get("Category")    or row.get("category")    or row.get("Kategori")    or row.get("kategori")    or ""
    description = row.get("Description") or row.get("description") or row.get("Beskrivning") or row.get("beskrivning") or ""
    return dish, price, category, description


# ---------------------------------------------------------------------------
# Evaluation
# ---------------------------------------------------------------------------

def evaluate_exact(scraped: list[dict], ground_truth: list[dict]) -> EvaluationResult:
    scraped_map: dict[str, list[dict]] = {}
    for row in scraped:
        dish, price, category, description = get_row_fields(row)
        key = normalize_name(dish)
        entry = {
            "original":         dish,
            "price":            normalize_price(price),
            "category":         normalize_category(category),
            "description":      description,
            "orig_price":       price,
            "orig_category":    category,
            "orig_description": description,
        }
        scraped_map.setdefault(key, []).append(entry)

    matched:         list[MatchedItem]       = []
    wrong_fields:    list[WrongFieldsItem]   = []
    missing:         list[str]               = []
    matched_indices: set[tuple[str, int]]    = set()

    for row in ground_truth:
        gt_dish, gt_price, gt_category, gt_description = get_row_fields(row)
        key           = normalize_name(gt_dish)
        norm_gt_price = normalize_price(gt_price)
        norm_gt_cat   = normalize_category(gt_category)

        scraped_key, scraped_idx = find_scraped_key(key, scraped_map, matched_indices)
        if scraped_key is None:
            missing.append(gt_dish)
            continue

        matched_indices.add((scraped_key, scraped_idx))
        s = scraped_map[scraped_key][scraped_idx]

        price_ok    = (not norm_gt_price) or (s["price"]    == norm_gt_price)
        category_ok = (not norm_gt_cat)   or category_similar(s["orig_category"], gt_category)
        desc_ok     = (not gt_description) or desc_similar(s["description"], gt_description)

        if price_ok and category_ok:
            matched.append(MatchedItem(
                dish=gt_dish,
                price=gt_price or s["orig_price"],
                category=gt_category or s["orig_category"],
                description=gt_description or s["orig_description"],
                desc_ok=desc_ok,
            ))
        else:
            wrong_fields.append(WrongFieldsItem(
                dish=gt_dish,
                price_ok=price_ok,
                category_ok=category_ok,
                desc_ok=desc_ok,
                scraped_price=s["orig_price"],
                expected_price=gt_price,
                scraped_category=s["orig_category"],
                expected_category=gt_category,
                scraped_description=s["orig_description"],
                expected_description=gt_description,
            ))

    hallucinated = [
        s["original"]
        for key, entries in scraped_map.items()
        for i, s in enumerate(entries)
        if (key, i) not in matched_indices
    ]

    n_matched = len(matched)
    n_wrong   = len(wrong_fields)
    n_hall    = len(hallucinated)
    n_miss    = len(missing)
    n_found   = n_matched + n_wrong

    precision = n_matched / (n_matched + n_wrong + n_hall) if (n_matched + n_wrong + n_hall) else 0.0
    recall    = n_matched / (n_matched + n_wrong + n_miss) if (n_matched + n_wrong + n_miss) else 0.0

    n_gt = n_found + n_miss
    name_accuracy = n_found / n_gt if n_gt else 0.0

    if n_found:
        price_accuracy    = (n_matched + sum(1 for i in wrong_fields if i.price_ok))    / n_found
        category_accuracy = (n_matched + sum(1 for i in wrong_fields if i.category_ok)) / n_found
        desc_accuracy     = (sum(1 for i in matched if i.desc_ok) + sum(1 for i in wrong_fields if i.desc_ok)) / n_found
    else:
        price_accuracy = category_accuracy = desc_accuracy = 0.0

    return EvaluationResult(
        matched=matched,
        wrong_fields=wrong_fields,
        missing=missing,
        hallucinated=hallucinated,
        precision=precision,
        recall=recall,
        name_accuracy=name_accuracy,
        price_accuracy=price_accuracy,
        category_accuracy=category_accuracy,
        description_accuracy=desc_accuracy,
    )


# ---------------------------------------------------------------------------
# Terminal report
# ---------------------------------------------------------------------------

GREEN  = ""
YELLOW = ""
RED    = ""
BOLD   = ""
RESET  = ""


def f1(precision: float, recall: float) -> float:
    return 2 * precision * recall / (precision + recall) if (precision + recall) else 0.0

def _color(value: float) -> str:
    if value >= 0.9: return GREEN
    if value >= 0.7: return YELLOW
    return RED


def print_report(result: EvaluationResult, restaurant: str) -> None:
    bar = "=" * 58
    print(f"\n{BOLD}{bar}{RESET}")
    print(f"{BOLD}  {restaurant.upper()}{RESET}")
    print(f"{BOLD}{bar}{RESET}")

    p, r = result.precision, result.recall
    f = f1(p, r)
    print(f"\n{BOLD}Huvud-metrics:{RESET}")
    print(f"  Precision   {_color(p)}{p:.1%}{RESET}")
    print(f"  Recall      {_color(r)}{r:.1%}{RESET}")
    print(f"  F1-score    {_color(f)}{f:.1%}{RESET}")

    na, pa, ca, da = result.name_accuracy, result.price_accuracy, result.category_accuracy, result.description_accuracy
    print(f"\n{BOLD}Per-fält:{RESET}")
    print(f"  Namn        {_color(na)}{na:.1%}{RESET}  (hittade av totalt GT)")
    print(f"  Pris        {_color(pa)}{pa:.1%}{RESET}  (av hittade)")
    print(f"  Kategori    {_color(ca)}{ca:.1%}{RESET}  (av hittade)")
    print(f"  Beskrivning {_color(da)}{da:.1%}{RESET}  (av hittade, bonus)")

    print(f"\n{BOLD}Sammanfattning:{RESET}")
    print(f"  {GREEN}✓ Rätt (namn+pris+kat)  {RESET}{len(result.matched)}")
    print(f"  {YELLOW}~ Fel pris/kategori     {RESET}{len(result.wrong_fields)}")
    print(f"  {RED}✗ Saknas                {RESET}{len(result.missing)}")
    print(f"  {RED}? Hallucinerade         {RESET}{len(result.hallucinated)}")

    if result.matched:
        print(f"\n{BOLD}{GREEN}✓ Korrekt matchade ({len(result.matched)}):{RESET}")
        for item in result.matched:
            desc_tag = "" if item.desc_ok else f"  {YELLOW}[beskr. avviker]{RESET}"
            print(f"    {item.dish:<45} {item.price}  [{item.category}]{desc_tag}")

    if result.wrong_fields:
        print(f"\n{BOLD}{YELLOW}~ Fel pris/kategori ({len(result.wrong_fields)}):{RESET}")
        for item in result.wrong_fields:
            print(f"    {item.dish}")
            if not item.price_ok:
                print(f"      pris:     scraped={item.scraped_price!r}  expected={item.expected_price!r}")
            if not item.category_ok:
                print(f"      kategori: scraped={item.scraped_category!r}  expected={item.expected_category!r}")

    if result.missing:
        print(f"\n{BOLD}{RED}✗ Saknas i scraped ({len(result.missing)}):{RESET}")
        for dish in result.missing:
            print(f"    {dish}")

    if result.hallucinated:
        print(f"\n{BOLD}{RED}? Hallucinerade ({len(result.hallucinated)}):{RESET}")
        for dish in result.hallucinated:
            print(f"    {dish}")


def print_aggregate_report(
    results: dict[str, EvaluationResult],
    missing_restaurants: list[str],
) -> None:
    bar = "=" * 58
    print(f"\n{BOLD}{bar}{RESET}")
    print(f"{BOLD}  AGGREGERAT — ALLA RESTAURANGER{RESET}")
    print(f"{BOLD}{bar}{RESET}")

    if missing_restaurants:
        print(f"\n{BOLD}{RED}Restauranger som saknas i scraped:{RESET}")
        for name in missing_restaurants:
            print(f"  {RED}✗ {name}{RESET}")

    # Micro-aggregate: pool all counts
    total_matched   = sum(len(r.matched)      for r in results.values())
    total_wrong     = sum(len(r.wrong_fields) for r in results.values())
    total_missing   = sum(len(r.missing)      for r in results.values())
    total_hall      = sum(len(r.hallucinated) for r in results.values())

    p = total_matched / (total_matched + total_wrong + total_hall) if (total_matched + total_wrong + total_hall) else 0.0
    r = total_matched / (total_matched + total_wrong + total_missing) if (total_matched + total_wrong + total_missing) else 0.0
    f = f1(p, r)

    print(f"\n{BOLD}Micro-aggregat  (alla rätter poolade):{RESET}")
    print(f"  Precision   {_color(p)}{p:.1%}{RESET}")
    print(f"  Recall      {_color(r)}{r:.1%}{RESET}")
    print(f"  F1-score    {_color(f)}{f:.1%}{RESET}")

    print(f"\n{BOLD}Totalt:{RESET}")
    print(f"  {GREEN}✓ Rätt        {RESET}{total_matched}")
    print(f"  {YELLOW}~ Fel fält    {RESET}{total_wrong}")
    print(f"  {RED}✗ Saknas      {RESET}{total_missing}")
    print(f"  {RED}? Hallucinerade {RESET}{total_hall}")

    print(f"\n{BOLD}Per restaurang (F1):{RESET}")
    for name, res in results.items():
        fi = f1(res.precision, res.recall)
        n_gt = len(res.matched) + len(res.wrong_fields) + len(res.missing)
        print(f"  {_color(fi)}{fi:.1%}{RESET}  {name}  ({n_gt} GT-rätter)")

    print(f"\n{BOLD}{bar}{RESET}\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Utvärdera meny-scraper (multi-sheet)")
    parser.add_argument("--scraped",      required=True, help="Skrapad Excel-fil med ett sheet per restaurang")
    parser.add_argument("--ground-truth", required=True, help="Ground truth Excel-fil med ett sheet per restaurang")
    parser.add_argument("--output", default="evaluation_report.json", help="Utdata JSON-rapport")
    args = parser.parse_args()

    print(f"Läser scraped:      {args.scraped}")
    scraped_sheets = read_xlsx_all_sheets(args.scraped)
    print(f"  {len(scraped_sheets)} sheets: {list(scraped_sheets)}")

    print(f"Läser ground truth: {args.ground_truth}")
    gt_sheets = read_xlsx_all_sheets(args.ground_truth)
    print(f"  {len(gt_sheets)} sheets: {list(gt_sheets)}")

    # Match sheets case-insensitively
    scraped_index = {name.lower().strip(): name for name in scraped_sheets}

    results: dict[str, EvaluationResult] = {}
    missing_restaurants: list[str] = []

    for gt_name, gt_rows in gt_sheets.items():
        scraped_key = scraped_index.get(gt_name.lower().strip())
        if scraped_key is None:
            print(f"  OBS: '{gt_name}' saknas i scraped-filen — räknas som 0% recall")
            missing_restaurants.append(gt_name)
            results[gt_name] = evaluate_exact([], gt_rows)
        else:
            results[gt_name] = evaluate_exact(scraped_sheets[scraped_key], gt_rows)

    for name, result in results.items():
        print_report(result, name)

    print_aggregate_report(results, missing_restaurants)

    report = {
        "restaurants": {
            name: {
                "matched":              [m.model_dump() for m in res.matched],
                "wrong_fields":         [w.model_dump() for w in res.wrong_fields],
                "missing":              res.missing,
                "hallucinated":         res.hallucinated,
                "precision":            res.precision,
                "recall":               res.recall,
                "f1":                   f1(res.precision, res.recall),
                "name_accuracy":        res.name_accuracy,
                "price_accuracy":       res.price_accuracy,
                "category_accuracy":    res.category_accuracy,
                "description_accuracy": res.description_accuracy,
            }
            for name, res in results.items()
        },
        "aggregate": {
            "precision": sum(len(r.matched) for r in results.values()) /
                         max(sum(len(r.matched) + len(r.wrong_fields) + len(r.hallucinated) for r in results.values()), 1),
            "recall":    sum(len(r.matched) for r in results.values()) /
                         max(sum(len(r.matched) + len(r.wrong_fields) + len(r.missing) for r in results.values()), 1),
        },
        "missing_restaurants": missing_restaurants,
    }
    report["aggregate"]["f1"] = f1(report["aggregate"]["precision"], report["aggregate"]["recall"])

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"Rapport sparad: {args.output}")


if __name__ == "__main__":
    main()
